import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_staff_csv(queryset):
    """
    Generates and returns a downloadable CSV response of staff records.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="staff_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Full Name', 'Email', 'Department', 'Job Title', 
        'Gender', 'Employment Type', 'Date Joined', 'Salary'
    ])

    for staff in queryset:
        user_name = staff.user.get_full_name() if staff.user else ""
        user_email = staff.user.email if staff.user else ""
        dept_name = staff.department.name if staff.department else ""

        writer.writerow([
            staff.id,
            user_name,
            user_email,
            dept_name,
            staff.job_title,
            staff.get_gender_display() if hasattr(staff, 'get_gender_display') else staff.gender,
            staff.get_employment_type_display() if hasattr(staff, 'get_employment_type_display') else staff.employment_type,
            staff.date_joined,
            f"{staff.salary:.2f}"
        ])

    return response


def export_staff_excel(queryset):
    """
    Generates and returns a professionally styled .xlsx workbook with summary statistics.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Directory"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="595959")
    data_font = Font(name=font_family, size=10)
    total_font = Font(name=font_family, size=11, bold=True)
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    total_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    # Title Header
    ws["A1"] = "Staff Management Report"
    ws["A1"].font = title_font
    ws["A2"] = "Exported Staff Directory & Salary Metrics"
    ws["A2"].font = subtitle_font

    # Table Headers
    headers = ["ID", "Full Name", "Email", "Department", "Job Title", "Gender", "Employment Type", "Date Joined", "Salary ($)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 6, 7, 8] else ("right" if col_idx == 9 else "left"), vertical="center")

    # Data Rows
    start_row = 5
    for row_idx, staff in enumerate(queryset, start=start_row):
        user_name = staff.user.get_full_name() if staff.user else ""
        user_email = staff.user.email if staff.user else ""
        dept_name = staff.department.name if staff.department else ""

        row_data = [
            staff.id, user_name, user_email, dept_name,
            staff.job_title, staff.gender, staff.employment_type,
            str(staff.date_joined), float(staff.salary or 0.00)
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = zebra_fill

            if col_idx in [1, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 9:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "$#,##0.00"

    # Total Row
    end_row = start_row + len(queryset) - 1
    if len(queryset) > 0:
        ws.cell(row=end_row + 1, column=1, value="Total").font = total_font
        total_cell = ws.cell(row=end_row + 1, column=9, value=f"=SUM(I{start_row}:I{end_row})")
        total_cell.font = total_font
        total_cell.number_format = "$#,##0.00"
        total_cell.alignment = Alignment(horizontal="right")

        for col in range(1, 10):
            ws.cell(row=end_row + 1, column=col).border = total_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # HTTP Response output
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="staff_report.xlsx"'
    wb.save(response)
    return response